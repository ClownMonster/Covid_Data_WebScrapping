'''
 Writing the sent data in the correct format to xlsx sheets
'''

import openpyxl

wb = ' '

wb = openpyxl.load_workbook('./CovidData.xlsx')

sheet = wb.active
r = 1
def set_header():
    global sheet
    # Setting the Headers
    sheet.cell(row = 1, column = 1).value = 'Country Name'
    sheet.cell(row = 1, column = 2).value = 'Total Cases'
    sheet.cell(row = 1, column = 3).value = 'New Cases'
    sheet.cell(row = 1, column = 4).value = 'Total Deaths Cases'
    sheet.cell(row = 1, column = 5).value = 'Total Recovered Cases'
    sheet.cell(row = 1, column = 6).value = 'Total Active Cases'
    sheet.cell(row = 1, column = 7).value = 'Critical Cases'
    sheet.cell(row = 1, column = 8).value = 'Total Cases Per 1M'
    sheet.cell(row = 1, column = 9).value = 'Total Deaths per 1M'
    sheet.cell(row = 1, column = 10).value = 'Total Tests Done'
    return

def write_the_data(Data_Recieved):
    global sheet, r
    r = r + 1
    for j in range(1,11):
        sheet.cell(row = r, column = j).value = Data_Recieved[j-1]

def save_close():
    global wb
    wb.save('CovidData.xlsx')
    wb.close()
    print('Saved the Workbook and Closed the session')